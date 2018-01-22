#!usr/bin/env python3.5

import openpyxl
"""
# opens excel file
wb = openpyxl.load_workbook('example.xlsx')

#The active sheet is the sheet that’s on top when the workbook is open.
active_sheet = wb.active

# get a certain cell value
#print(active_sheet["A1"].value)
cell = active_sheet["A1"]
cell_one  = 'Row ' + str(cell.row) + ', Column ' + cell.column + ' is ' + cell.value
#print(cell_one)



#specify by row and column number

my_value = active_sheet.cell(row=1, column=2)

# get rows and columns from sheets

all_sheets = tuple(active_sheet['A1':'C3'])

for row in all_sheets:
 for cell in row: 
 	print(cell.coordinate, cell.value)
"""
# Writing to Excel

wb = openpyxl.Workbook()
print(wb.get_sheet_names())
ws = wb.create_sheet()
ws.title = 'Spam Bacon Eggs Sheet'
print(wb.get_sheet_names())

